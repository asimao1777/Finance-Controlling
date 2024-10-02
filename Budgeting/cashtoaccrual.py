# Description: the code takes an Excel workbook with the budget database, creates a new column and
#              converts the cash value column into its accrual value, placing the value converted
#              into the newly created column. The conversion is based on the days of payable,
#              which the user shall define in the code.


import openpyxl
import pandas as pd

option = str(input(f'Do you want to update days payable/receivable of any category? Yes or No? :  ')).upper()

if option == "YES":
    sales = int(input(f'Insert the average days of receivables for sales:  '))
    services = int(input(f'Insert the average days of payables for services:  '))
    eng_services = int(input(f'Insert the average days of payables for engineering services:  '))
    equipment = int(input(f'Insert the average days of payables for equipment:  '))
    supplies = int(input(f'Insert the average days of payables for supplies:  '))
    travel = int(input(f'Insert the average days of payables for travel:  '))
    salaries = int(input(f'Insert the average days of payables for salaries:  '))
    social_charges = int(input(f'Insert the average days of payables for social charges:  '))
    construction = int(input(f'Insert the average days of payables for construction:  '))
    other = int(input(f'Insert the average days of payables for other:  '))

else:
    sales = 45
    services = 30
    eng_services = 60
    equipment = 90
    supplies = 35
    travel = 30
    salaries = 30
    social_charges = 45
    construction = 90
    other = 45

days_of_payables = {'sales': sales,
                    'services': services,
                    'engineering services': eng_services,
                    'equipment': equipment,
                    'supplies': supplies,
                    'travel': travel,
                    'salaries': salaries,
                    'social charges': social_charges,
                    'construction': construction,
                    'other': other}

print(days_of_payables)

days_accum = {'Jan': 31,
              'Feb': 59,
              'Mar': 90,
              'Apr': 120,
              'May': 151,
              'Jun': 181,
              'Jul': 212,
              'Aug': 243,
              'Sep': 273,
              'Oct': 304,
              'Nov': 334,
              'Dec': 365}

prev_year = 2023
curr_year = 2024

db = input(f'Please type the address and name of the database file to be uploaded: ')
ws = openpyxl.load_workbook(db)

sheet = input(f'Please type the name of the sheet of database workbook to be active, if any: ')
aba = ws[sheet]

for linha in range(2, aba.max_row + 1):

    month = aba.cell(row=linha, column=2).value
    days_acc = days_accum[month]
    #     print(days_acc)

    pay_cat = aba.cell(row=linha, column=5).value
    days_cat = days_of_payables[pay_cat]
    #     print(days_cat)

    nro_days = days_acc - days_cat
    #     print(total)

    if nro_days <= days_accum['Jan']:
        aba.cell(row=linha, column=11).value = 'Prev Years'
        aba.cell(row=linha, column=12).value = prev_year

    elif nro_days <= days_accum['Feb']:
        aba.cell(row=linha, column=11).value = 'Jan'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Mar']:
        aba.cell(row=linha, column=11).value = 'Feb'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Apr']:
        aba.cell(row=linha, column=11).value = 'Mar'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['May']:
        aba.cell(row=linha, column=11).value = 'Apr'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Jun']:
        aba.cell(row=linha, column=11).value = 'May'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Jul']:
        aba.cell(row=linha, column=11).value = 'Jun'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Aug']:
        aba.cell(row=linha, column=11).value = 'Jul'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Sep']:
        aba.cell(row=linha, column=11).value = 'Aug'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Oct']:
        aba.cell(row=linha, column=11).value = 'Sep'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Nov']:
        aba.cell(row=linha, column=11).value = 'Oct'
        aba.cell(row=linha, column=12).value = curr_year

    elif nro_days <= days_accum['Dec']:
        aba.cell(row=linha, column=11).value = 'Nov'
        aba.cell(row=linha, column=12).value = curr_year

    else:
        aba.cell(row=linha, column=11).value = 'Dec'
        aba.cell(row=linha, column=12).value = curr_year

ws.save(db)
